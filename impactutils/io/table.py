# stdlib imports
import json
import re
import string
import time

# third party imports
import pandas as pd
import numpy as np
from lxml import etree
from openpyxl import load_workbook, utils

REQUIRED_COLUMNS = ['STATION', 'LAT', 'LON', 'NETID']
CHANNEL_GROUPS = [['[A-Z]{2}E', '[A-Z]{2}N', '[A-Z]{2}Z'],
                  ['[A-Z]{2}1', '[A-Z]{2}2', '[A-Z]{2}Z'],
                  ['H1', 'H2', 'Z'],
                  ['UNK']]
COMPONENTS = ['GREATER_OF_TWO_HORIZONTALS',
              'GEOMETRIC_MEAN', 'ARITHMETIC MEAN']
CHANNEL_PATTERNS = ['^[H,B][H,L,N][E,N,Z,1,2,3]$',  # match standard seed names
                    '^H[1,2]$',  # match H1/H2
                    '^Z$']  # match Z
PGM_COLS = ['PGA', 'PGV', 'SA(0.3)', 'SA(1.0)', 'SA(3.0)']
OPTIONAL = ['NAME', 'DISTANCE', 'REFERENCE',
            'INTENSITY', 'SOURCE', 'LOC', 'INSTTYPE', 'ELEV',
            'NRESP', 'INTENSITY_STDDEV', '', 'FLAG', 'INSTRUMENT', 'PERIOD',
            'SENSITIVITY', 'SERIAL', 'SOURCE_FORMAT', 'STRUCTURE', 'DAMPING']
FLOATRE = "[-+]?[0-9]*\.?[0-9]+"


def dataframe_to_json(df, jsonfile, duration_interval="5-95", arias_units="m/s",
                      duration_units="s", fas_units="cm/s", pga_units="%g", pgv_units="cm/s",
                      sa_units="%g"):
    """Write a dataframe to json format.

    Args:
        df (DataFrame): Pandas dataframe, as described in read_excel.
        jsonfile (str): Path to file where json file should be written.
        duration_interval (str): Interval over which duration values were computed.
        arias_units (str): Override default arias units (m/s).
        duration_units (str): Override default duration units (s).
        fas_units (str): Override default fas units (cm/s).
        pga_units (str): Override default pga units (%g).
        pgv_units (str): Override default pgv units (cm/s).
        sa_units (str): Override default sa units (%g).
    Notes:
        This method accepts either a dataframe from read_excel, or
        one with this structure:
         - STATION: Station code (REQUIRED)
         - IMC: Component (HHE,HHN, GREATER_OF_TWO, ROTD50, etc.) (REQUIRED)
         - IMT: Intensity measure type (pga,pgv, etc.) (REQUIRED)
         - VALUE: IMT value. (REQUIRED)
         - LAT: Station latitude. (REQUIRED)
         - LON: Station longitude. (REQUIRED)
         - NETID: Station contributing network. (REQUIRED)
         - FLAG: String quality flag, meaningful to contributing networks,
                 but ShakeMap ignores any station with a non-zero value. (REQUIRED)
         - ELEV: Elevation of station (m). (OPTIONAL)
         - NAME: String describing station. (OPTIONAL)
         - DISTANCE: Distance (km) from station to origin. (OPTIONAL)
         - LOC: Description of location (i.e., "5 km south of Wellington")
                (OPTIONAL)
         - INSTTYPE: Instrument type (FBA, etc.) (OPTIONAL)
         - INTENSITY: MMI intensity. (OPTIONAL)
         - NRESP: Number of responses for aggregated intensity. (OPTIONAL)
         - INTENSITY_STDDEV: Uncertainty for this intensity. (OPTIONAL)
         - INSTRUMENT: Type of instrument. (OPTIONAL)
         - SERIAL: Instrument serial number. (OPTIONAL)
         - PERIOD: Sampling period. (OPTIONAL)
         - DAMPING: Damping. (OPTIONAL)
         - SENSITIVITY: Instrument sensitivity. (OPTIONAL)
         - SOURCE_FORMAT: Format of source file. (OPTIONAL)
         - STRUCTURE: Type of str. (OPTIONAL)
    """
    features = []
    units = {
        'ARIAS': arias_units,
        'DURATION': duration_units,
        'FAS': fas_units,
        'PGA': pga_units,
        'PGV': pgv_units,
        'SA': sa_units
    }

    if hasattr(df.columns, 'levels'):
        top_headers = set(df.columns.levels[0])
        possible_columns = ['NAME', 'DISTANCE', 'INTENSITY', 'NETID',
                            'STATION', 'LAT', 'LON', 'ELEV',
                            'INTENSITY', 'NRESP', 'STRUCTURE', 'SOURCE_FORMAT',
                            'SENSITIVITY', 'DAMPING', 'PERIOD', 'SERIAL',
                            'INSTRUMENT', 'INSTTYPE', 'LOC', 'SOURCE',
                            'INTENSITY_STDDEV', 'FLAG']
        channels = (top_headers - set(possible_columns))
    else:
        channels = []
    processed_stations = []
    for _, row in df.iterrows():
        tmprow = row.copy()
        if isinstance(tmprow.index, pd.core.indexes.multi.MultiIndex):
            tmprow.index = tmprow.index.droplevel(1)

        # Properties\
        stationcode, flag, props = _get_properties(tmprow)

        # Setup station dictionary
        if stationcode in processed_stations:
            continue
        station = {}
        station['id'] = stationcode
        station['type'] = 'Feature'
        station['properties'] = props

        # Geometry
        geom = {}
        geom['type'] = 'Point'
        geom['coordinates'] = [
            tmprow['LAT'], tmprow['LON'], tmprow['ELEV']
        ]
        station['geometry'] = geom

        # sort channels
        if 'imt' not in tmprow.index and 'IMT' not in tmprow.index:
            channels = sorted(list(channels))
            if len(channels) != len(set(channels)):
                msg = f'IMC types must be unique.'
                raise Exception(msg)
            # Create channels section if necessary
            for channel in channels:
                # Get whether it is a component or a channel
                checked_channel = _get_channels([channel])
                if len(checked_channel) == 0:
                    channel_type = 'component'
                    if 'components' not in station['properties']:
                        station['properties']['components'] = {}
                else:
                    channel_type = 'channel'
                    channel = checked_channel[0]
                    if 'channels' not in station['properties']:
                        station['properties']['channels'] = []
                imts = row[channel].index.values
                if len(imts) != len(set(imts)):
                    msg = f'IMTS under each IMC (e.g. {channel}) must be unique.'
                    raise Exception(msg)
                # DEAL WITH OLD VERSION
                # DEAL WITH OLD VERSION
                if channel_type == 'channel':
                    imc_dict = {
                        'name': channel.upper(),
                        'amplitudes': []
                    }
                    # lowercase for the old version
                    for imt in imts:
                        if _empty_row(row[channel][imt]):
                            continue
                        imt_dict = _set_channel_props(duration_interval, flag,
                                                      imt, units, row[channel][imt])
                        imc_dict['amplitudes'] += [imt_dict]
                    station['properties']['channels'] += [imc_dict]
                # DEAL WITH NEW VERSION
                if channel_type == 'component':
                    station['properties']['components'][channel.upper()] = {}
                    for imt in imts:
                        if _empty_row(row[channel][imt]):
                            continue
                        value = row[channel][imt]
                        station = _set_component_props(channel, duration_interval,
                                                       flag, imt, station, units, value)
            features += [station]
        else:
            # this file was created by a process that has imt/value columns
            # search the dataframe for all rows with this same station code
            scode = tmprow['STATION']
            station_rows = df[df['STATION'] == scode]
            channels = station_rows['CHANNEL'].unique()
            for channel in channels:
                # Get whether it is a component or a channel
                checked_channel = _get_channels([channel])
                if len(checked_channel) == 0:
                    channel_type = 'component'
                    if 'components' not in station['properties']:
                        station['properties']['components'] = {}
                else:
                    channel_type = 'channel'
                    channel = checked_channel[0]
                    if 'channels' not in station['properties']:
                        station['properties']['channels'] = []
                rows = station_rows[station_rows['CHANNEL'] == channel]
                imts = rows['IMT'].values
                if len(imts) != len(set(imts)):
                    msg = f'IMTS under each IMC (e.g. {channel}) must be unique.'
                    raise Exception(msg)
                # DEAL WITH OLD VERSION
                if channel_type == 'channel':
                    imc_dict = {
                        'name': channel.upper(),
                        'amplitudes': []
                    }
                    # lowercase for the old version
                    for imt in imts:
                        row = rows[rows['IMT'] == imt]
                        if _empty_row(row['VALUE'].values[0]):
                            continue
                        imt_dict = _set_channel_props(duration_interval,
                                                      row['FLAG'].values[0], imt, units,
                                                      row['VALUE'].values[0])
                        imc_dict['amplitudes'] += [imt_dict]
                    station['properties']['channels'] += [imc_dict]
                # DEAL WITH NEW VERSION
                if channel_type == 'component':
                    station['properties']['components'][channel.upper()] = {}
                    for imt in imts:
                        row = rows[rows['IMT'] == imt]
                        if _empty_row(row['VALUE'].values[0]):
                            continue
                        value = row['VALUE'].values[0]
                        station = _set_component_props(channel, duration_interval,
                                                       row['FLAG'].values[0], imt, station,
                                                       units, value)
            features += [station]
        processed_stations.append(stationcode)
    geojson = {
        'type': "FeatureCollection",
        'features': features
    }
    return geojson


def dataframe_to_xml(df, xmlfile, reference=None):
    """Write a dataframe to ShakeMap XML format.

    This method accepts either a dataframe from read_excel, or
    one with this structure:
     - STATION: Station code (REQUIRED)
     - CHANNEL: Channel (HHE,HHN, etc.) (REQUIRED)
     - IMT: Intensity measure type (pga,pgv, etc.) (REQUIRED)
     - VALUE: IMT value. (REQUIRED)
     - LAT: Station latitude. (REQUIRED)
     - LON: Station longitude. (REQUIRED)
     - NETID: Station contributing network. (REQUIRED)
     - FLAG: String quality flag, meaningful to contributing networks,
             but ShakeMap ignores any station with a non-zero value. (REQUIRED)
     - ELEV: Elevation of station (m). (OPTIONAL)
     - NAME: String describing station. (OPTIONAL)
     - DISTANCE: Distance (km) from station to origin. (OPTIONAL)
     - LOC: Description of location (i.e., "5 km south of Wellington")
            (OPTIONAL)
     - INSTTYPE: Instrument type (FBA, etc.) (OPTIONAL)
     - INTENSITY: MMI intensity. (OPTIONAL)
     - NRESP: Number of responses for aggregated intensity. (OPTIONAL)
     - INTENSITY_STDDEV: Uncertainty for this intensity. (OPTIONAL)

    Args:
        df (DataFrame): Pandas dataframe, as described in read_excel.
        xmlfile (str): Path to file where XML file should be written.
    """
    if hasattr(df.columns, 'levels'):
        top_headers = set(df.columns.levels[0])
        required = set(REQUIRED_COLUMNS)
        optional = set(OPTIONAL)
        channel_candidates = (top_headers - required) - optional
        channels = _get_channels(channel_candidates)
    else:
        channels = []
    root = etree.Element(
        'shakemap-data', code_version="3.5", map_version="3")

    create_time = int(time.time())
    stationlist = etree.SubElement(
        root, 'stationlist', created=f'{int(create_time):d}')
    if reference is not None:
        stationlist.attrib['reference'] = reference

    processed_stations = []

    for _, row in df.iterrows():
        tmprow = row.copy()
        if isinstance(tmprow.index, pd.core.indexes.multi.MultiIndex):
            tmprow.index = tmprow.index.droplevel(1)

        # assign required columns
        stationcode = str(tmprow['STATION']).strip()

        netid = tmprow['NETID'].strip()
        if not stationcode.startswith(netid):
            stationcode = f'{netid}.{stationcode}'

        # if this is a dataframe created by shakemap,
        # there will be multiple rows per station.
        # below we process all those rows at once,
        # so we need this bookkeeping to know that
        # we've already dealt with this station
        if stationcode in processed_stations:
            continue

        station = etree.SubElement(stationlist, 'station')

        station.attrib['code'] = stationcode
        station.attrib['lat'] = f"{tmprow['LAT']:.4f}"
        station.attrib['lon'] = f"{tmprow['LON']:.4f}"

        # assign optional columns
        if 'NAME' in tmprow:
            station.attrib['name'] = tmprow['NAME'].strip()
        if 'NETID' in tmprow:
            station.attrib['netid'] = tmprow['NETID'].strip()
        if 'DISTANCE' in tmprow:
            station.attrib['dist'] = f"{tmprow['DISTANCE']:.1f}"
        if 'INTENSITY' in tmprow:
            station.attrib['intensity'] = f"{tmprow['INTENSITY']:.1f}"
        if 'NRESP' in tmprow:
            station.attrib['nresp'] = f"{int(tmprow['NRESP']):d}"
        if 'INTENSITY_STDDEV' in tmprow:
            station.attrib['intensity_stddev'] = f"{tmprow['INTENSITY_STDDEV']:.2f}"
        if 'SOURCE' in tmprow:
            station.attrib['source'] = tmprow['SOURCE'].strip()
        if 'LOC' in tmprow:
            station.attrib['loc'] = tmprow['LOC'].strip()
        if 'INSTTYPE' in tmprow:
            station.attrib['insttype'] = tmprow['INSTTYPE'].strip()
        if 'ELEV' in tmprow:
            station.attrib['elev'] = f"{tmprow['ELEV']:.1f}"

        if 'imt' not in tmprow.index:
            # sort channels by N,E,Z or H1,H2,Z
            channels = sorted(list(channels))

            for channel in channels:
                component = etree.SubElement(station, 'comp')
                component.attrib['name'] = channel.upper()

                # figure out if channel is horizontal or vertical
                if channel[-1] in ['1', '2', 'E', 'N']:
                    component.attrib['orientation'] = 'h'
                else:
                    component.attrib['orientation'] = 'z'

                # create sub elements out of any of the PGMs
                # this is extra confusing because we're trying to
                # transition from psa03 style to SA(0.3) style.
                # station xml format only accepts the former, but we're
                # supporting the latter as input, and the format as output.

                # loop over desired output fields
                for pgm in ['pga', 'pgv', 'psa03', 'psa10', 'psa30']:
                    newpgm = _translate_imt(pgm, row[channel].index)
                    c1 = newpgm not in row[channel]
                    c2 = False
                    if not c1:
                        c2 = np.isnan(row[channel][newpgm])
                    if c1 or c2:
                        continue
                    # make an element with the old style name
                    pgm_el = etree.SubElement(component, pgm)
                    pgm_el.attrib['flag'] = '0'
                    pgm_el.attrib['value'] = f'{row[channel][newpgm]:.4f}'
            processed_stations.append(stationcode)
        else:
            # this file was created by a process that has imt/value columns
            # search the dataframe for all rows with this same station code
            scode = tmprow['STATION']
            station_rows = df[df['STATION'] == scode]

            # now we need to find all of the channels
            channels = station_rows['channel'].unique()
            for channel in channels:
                channel_rows = station_rows[station_rows['channel'] == channel]
                component = etree.SubElement(station, 'comp')
                component.attrib['name'] = channel.upper()
                for _, channel_row in channel_rows.iterrows():
                    pgm = channel_row['imt']
                    value = channel_row['value']

                    pgm_el = etree.SubElement(component, pgm)
                    pgm_el.attrib['value'] = f'{value:.4f}'
                    pgm_el.attrib['flag'] = str(channel_row['flag'])

            processed_stations.append(stationcode)

    tree = etree.ElementTree(root)
    tree.write(xmlfile, pretty_print=True)


def read_excel(excelfile):
    """Read strong motion Excel spreadsheet, return a DataFrame.

    Args:
        excelfile (str): Path to valid Excel file.
    Returns:
        DataFrame: Multi-indexed dataframe as described below.
        str or None: Reference string or None.

     - "STATION" String containing UNIQUE identifying station information.
     - "LAT" Latitude where peak ground motion observations were made.
     - "LON" Longitude where peak ground motion observations were made.
     - "NETID" (usually) two letter code indicating the source network.

    Optional columns include:
     - "NAME" String describing area where peak ground motions were observed.
     - "SOURCE" String describing (usu. long form) source of peak ground
       motion data.
     - "DISTANCE" Distance from epicenter to station location, in units of km.
     - "LOC" Two character location code.
     - "INSTTYPE" Instrument type, str.
     - "ELEV" Station elevation, in meters.

    And then at least one of the following columns:
     - "INTENSITY" MMI value (1-10).

        AND/OR
      a grouped set of per-channel peak ground motion columns, like this:

      -------------------------------------------------------------------------------
      |         H1              |           H2            |             Z           |
      -------------------------------------------------------------------------------
      |pga|pgv|psa03|psa10|psa30|pga|pgv|psa03|psa10|psa30|pga|pgv|psa03|psa10|psa30|
      -------------------------------------------------------------------------------

      The peak ground motion columns can be any of the following:
      - "PGA" Peak ground acceleration in units of %g.
      - "PGV" Peak ground velocity in units of cm/sec.
      - "PSA03" Peak spectral acceleration at 0.3 seconds, in units of %g.
      - "PSA10" Peak spectral acceleration at 1.0 seconds, in units of %g.
      - "PSA30" Peak spectral acceleration at 3.0 seconds, in units of %g.

    Valid "channel" columns are {H1,H2,Z} or {XXN,XXE,XXZ}, where 'XX' is any
    two-letter combination, usually adhering to the following standard:
    http://www.fdsn.org/seed_manual/SEEDManual_V2.4_Appendix-A.pdf

    If the input data set provides no channel information, then the channel
    can be simply "UNK".

    """

    # figure out if data frame is multi-index or not
    wb = load_workbook(excelfile)
    ws = wb.active

    # figure out where the top left of the data begins
    topleft = 'A1'
    first_cell = 'A2'
    second_cell = 'A3'

    # figure out if there is a little reference section in this...
    reference = None
    skip_rows = None
    header = [0, 1]
    if ws[topleft].value.lower() != 'reference':
        raise KeyError('Reference cells are required in A1 and B1!')
    refcell = _move(topleft, 0, 1)
    reference = ws[refcell].value
    first_cell = _move(topleft, 1, 0)
    second_cell = _move(first_cell, 1, 0)
    skip_rows = [0]
    header = [1, 2]

    is_multi = True
    # if the first column of the second row is not empty,
    # then we do not have a multi-index.
    if ws[second_cell].value is not None:
        is_multi = False

    # read in dataframe, assuming that ground motions are grouped by channel
    if is_multi:
        try:
            # note - in versions of pandas prior to 0.24, index_col=None
            # has no effect here.  Hence the unsetting of the index later.
            df = pd.read_excel(excelfile, header=header, index_col=None)
            # if the name column is all blanks, it's filled with NaNs by
            # default, which causes problems later on.  Replace with
            # empty strings
            if 'NAME' in df.columns:
                df['NAME'] = df['NAME'].fillna('')
        except pd.errors.ParserError:
            raise IndexError('Input file has invalid empty first data row.')

        headers = df.columns.get_level_values(0).str.upper()
        subheaders = df.columns.get_level_values(1).str.upper()
        df.columns = pd.MultiIndex.from_arrays([headers, subheaders])
        top_headers = df.columns.levels[0]
    else:
        df = pd.read_excel(excelfile, skiprows=skip_rows, index_col=None)
        top_headers = df.columns

    # make sure basic columns are present
    if 'STATION' not in top_headers:
        df['STATION'] = df.index
        df = df.reset_index(drop=True)
        top_headers = df.columns.levels[0]
    if not set(REQUIRED_COLUMNS).issubset(set(top_headers)):
        fmt = f'Input Excel file must specify the following columns: {(str(REQUIRED_COLUMNS))}.'
        raise KeyError(fmt)

    # check if channel headers are valid
    channels = (set(top_headers) - set(REQUIRED_COLUMNS)) - set(OPTIONAL)
    valid = False
    if len(channels):
        channel_copy = [c for c in channels
                        if c.upper() not in COMPONENTS and not c.upper().startswith('ROT')]
        for channel_group in CHANNEL_GROUPS:
            num_channels = 0
            for channel_pat in channel_group:
                cp = re.compile(channel_pat)
                if len(list(filter(cp.match, channel_copy))):
                    num_channels += 1
            if num_channels == 1 and len(channel_copy) == 1:
                valid = True
                break
            elif num_channels > 1:
                h1_pat = re.compile(channel_group[0])
                h2_pat = re.compile(channel_group[1])
                has_h1 = len(list(filter(h1_pat.match, channel_copy))) > 0
                has_h2 = len(list(filter(h2_pat.match, channel_copy))) > 0
                if has_h1 or has_h2:
                    valid = True
                    break
    else:
        valid = True
    if not valid:
        raise KeyError(
            f'{str(sorted(list(channels)))} is not a valid channel grouping')

    # make sure the empty cells are all nans or floats
    found = False
    if 'INTENSITY' in top_headers:
        found = True
    empty_cell = re.compile(r'\s+')
    for channel in channels:
        channel_df = df[channel].copy()
        for column in PGM_COLS:
            if column in channel_df:
                found = True
                channel_df[column] = channel_df[column].replace(
                    empty_cell, np.nan)
                channel_df[column] = channel_df[column].astype(float)
        df[channel] = channel_df

    if not found:
        intensity_col = str(PGM_COLS + ['intensity'])
        fmt = (f'File must contain at least one of the following '
               f'data columns: {intensity_col}')
        raise KeyError(fmt)

    return (df, reference)


def _empty_row(value):
    """Check weather the imt row is empty

    Args:
        value (float or int): The cell value.

    Returns:
        bool: Whether or not the cell is empty.
    """
    val = str(value).lower()
    if val == 'nan' or val == 'null':
        return True
    else:
        return False


def _get_channels(columns):
    channels = []
    for column in columns:
        for cmatch in CHANNEL_PATTERNS:
            if re.search(cmatch, column) is not None:
                channels.append(column)
                break
    return channels


def _get_period(imt):
    """Extract the period from the imt string.

    Args:
        imt (str): The imt type.

    Returns:
        float: Period over which fas or sa were computed.
    """
    if imt.lower().startswith('psa'):
        if imt.endswith('03'):
            period = 0.3
        elif imt.endswith('02'):
            period = 0.2
        elif imt.endswith('01'):
            period = 0.1
        elif imt.endswith('10'):
            period = 1.0
        elif imt.endswith('20'):
            period = 2.0
        elif imt.endswith('30'):
            period = 3.0
    else:
        period = float(re.search(FLOATRE, imt).group())
    return period


def _get_properties(df):
    """Get the properties for the table

    Args:
        df (pandas.Dataframe): Dataframe with the property columns.

    Returns:
        (str, int, dict): 'stationcode' formatted string; 'flag' integer for
            the station; dictionary of 'props' (properties) for the station.
    """
    props = {}
    required = ['STATION', 'NETID', 'FLAG', 'LAT', 'LON', 'ELEV']
    if len(set(required).difference(df.index.values)) > 0:
        msg = f'Missing one of the required columns: {required}.'
        raise Exception(msg)

    # Required property columns
    flag = df['FLAG']
    props['network'] = df['NETID'].strip()
    code = str(df['STATION']).strip()
    netid = df['NETID'].strip()
    if not code.startswith(netid):
        stationcode = f'{netid}.{code}'
    props['code'] = code
    props['station'] = code

    # Old Standard
    if 'NAME' in df:
        props['name'] = df['NAME'].strip()
    if 'DISTANCE' in df:
        props['distance'] = df['DISTANCE']
    if 'INTENSITY' in df:
        props['intensity'] = df['INTENSITY']
    if 'NRESP' in df:
        props['nresp'] = int(df['NRESP'])
    if 'INTENSITY_STDDEV' in df:
        props['intensity_stddev'] = df['INTENSITY_STDDEV']
    if 'SOURCE' in df:
        props['provider'] = df['SOURCE'].strip()
    if 'LOC' in df:
        loc = df['LOC'].strip()
        if loc == '':
            loc == '--'
        props['location'] = loc
    else:
        props['location'] = '--'
    if 'INSTTYPE' in df:
        props['type'] = df['INSTTYPE'].strip()
    # new format properties
    if 'INSTRUMENT' in df:
        props['instrument'] = df['INSTRUMENT'].strip()
    if 'SERIAL' in df:
        props['serial'] = df['SERIAL']
    else:
        props['serial'] = 'None'
    if 'PERIOD' in df:
        props['period'] = df['PERIOD']
    if 'DAMPING' in df:
        props['damping'] = df['DAMPING']
    if 'SENSITIVITY' in df:
        props['sensitivity'] = df['SENSITIVITY']
    if 'SOURCE_FORMAT' in df:
        props['source_format'] = df['SOURCE_FORMAT'].strip()
    if 'STRUCTURE' in df:
        props['structure'] = df['STRUCTURE']
    return stationcode, flag, props


def _get_units(imt, units):
    """Determine the units for the current imt.

    Args:
        imt (str): The imt type.
        units (dict): The dictionary defining the units.

    Returns:
        str: Units associated with the imt.
    """
    imt = imt.upper()
    if imt.startswith('SA') or imt.startswith('PSA'):
        unit = units['SA']
    elif imt.startswith('FAS'):
        unit = units['FAS']
    else:
        if imt in units:
            unit = units[imt]
        else:
            unit = 'unknown'
    return unit


def _move(cellstr, nrows, ncols):
    """Internal method for adding rows/columns to cell coordinate.

    'A1' moved by 1 row, 1 column => 'B2'

    Args:
        cellstr (str): Cell coordinate (A1,B2)
        nrows (int): Number of rows to move (usually down)
        ncols (int): Number of columns to move (usually right)
    Returns:
        str: New cell coordinate.
    """
    # WARNING! This will only work up to column Z!
    # colidx is a string, rowidx is a number
    col_str_idx, rowidx = utils.cell.coordinate_from_string(cellstr)
    letters = string.ascii_uppercase
    try:
        colidx = letters.index(col_str_idx)
        newcolidx = colidx + ncols
        newrowidx = rowidx + nrows
        newcellstr = f'{letters[newcolidx]}{int(newrowidx):d}'
        return newcellstr
    except ValueError:
        raise ValueError(
            f'Could not add {int(ncols):d} columns to column {col_str_idx}.')


def _set_channel_props(duration_interval, flag, imt, units, value):
    """Set all of the values in the channel dictionary.

    Args:
        duration_interval (str): The interval over which the duration was computed.
        flag (int): The flag associated with
        imt (str): The imt type.
        units (dict): The dictionary defining the units.
        value (float): Column value.

    Returns:
        dict: dictionary of the imt.
    """
    imt_dict = {}
    imt_dict['flag'] = flag
    imt_dict['ln_sigma'] = 0
    imt_dict['value'] = value
    imt_dict['units'] = _get_units(imt.upper(), units)
    imt_dict = _set_channel_props = (imt_dict)
    if imt.lower().startswith('sa') or imt.lower().startswith('psa'):
        period = _get_period(imt)
        imt_dict['name'] = f'sa({period})'
    elif imt.lower().startswith('fas'):
        period = float(re.search(FLOATRE, imt).group())
        imt_dict['name'] = f'fas({period})'
    elif imt.lower().startswith('duration'):
        imt_dict['name'] = 'duration'
        imt_dict['interval'] = duration_interval
    else:
        imt_dict['name'] = imt.lower()
    return imt_dict


def _set_component_props(channel, duration_interval, flag, imt, station, units, value):
    """Add values (component dictionary) to the station dictionary.

    Args:
        channel (str): The channel type.
        duration_interval (str): The interval over which the duration was computed.
        flag (int): The flag associated with
        imt (str): The imt type.
        station (dict): The station dictionary.
        units (dict): The dictionary defining the units.
        value (float): Column value.

    Returns:
        dict: dictionary of the station.
    """
    imt_dict = {}
    imt_dict['flag'] = flag
    imt_dict['ln_sigma'] = 0
    imt_dict['units'] = _get_units(imt.upper(), units)
    imt_dict['value'] = value
    if imt.lower().startswith('sa') or imt.lower().startswith('psa'):
        if 'SA' not in station['properties']['components'][channel.upper()]:
            station['properties']['components'][channel.upper()]['SA'] = [
            ]
        imt_dict['period'] = _get_period(imt)
        station['properties']['components'][channel.upper()
                                            ]['SA'] += [imt_dict]
    elif imt.lower().startswith('fas'):
        if 'FAS' not in station['properties']['components'][channel.upper()]:
            station['properties']['components'][channel.upper()]['FAS'] = [
            ]
        imt_dict['period'] = float(
            re.search(FLOATRE, imt).group())
        station['properties']['components'][channel.upper()
                                            ]['FAS'] += [imt_dict]
    elif imt.lower().startswith('duration'):
        imt_dict['interval'] = duration_interval
        station['properties']['components'][channel.upper()
                                            ]['DURATION'] = imt_dict
    else:
        station['properties']['components']
        imt_str = imt.upper()
        station['properties']['components'][channel.upper()
                                            ][imt_str] = imt_dict
    return station


def _translate_imt(oldimt, imtlist):
    # translate from psa03 to sa(0.3)
    if oldimt.upper() in ['PGA', 'PGV']:
        newimt = oldimt.upper()
    else:
        match = re.search(r'\d+', oldimt)
        if match is not None:
            period = float(match.group()) / 10
            for imt in imtlist:
                if not imt.startswith('SA'):
                    continue
                try:
                    imt_period = float(re.search(FLOATRE, imt).group())
                except Exception:
                    continue
                if imt_period == period:
                    newimt = imt
                    break
        else:
            newimt = ''
    return newimt
